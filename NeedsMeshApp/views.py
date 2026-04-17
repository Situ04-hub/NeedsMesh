import csv
import io
import json
from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import transaction
from django.core.mail import send_mail
from django.conf import settings
from .models import (
    UserProfile, Locality, CommunityProblem, ProofImage,
    SurveySubmission, Notification, CATEGORY_CHOICES, Organisation,
    ProblemMessage, PasswordResetOTP
)
from .forms import (
    RegistrationForm, LoginForm, ProfileForm, SurveyForm,
    ProofImageForm, FinaliseEventForm, ResolveProblemForm,
    KillProblemForm, LocalityForm, CSVImportForm, SetVolunteersRequiredForm
)

# ============================================================
# Helper utilities
# ============================================================

def role_required(*roles):
    """Decorator – redirect non-matching roles with a message."""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            try:
                profile = request.user.profile
            except UserProfile.DoesNotExist:
                messages.error(request, "Your profile is not set up. Contact admin.")
                return redirect('login')
            if profile.role not in roles:
                messages.error(request, "You do not have permission to access that page.")
                return redirect('dashboard')
            return view_func(request, *args, **kwargs)
        wrapper.__name__ = view_func.__name__
        return wrapper
    return decorator


def send_notification(recipient, message, link=''):
    """Create an in-app notification for a user."""
    Notification.objects.create(recipient=recipient, message=message, link=link)


def send_email_notification(user, subject, body):
    """Send email; silently fails in dev (console backend)."""
    try:
        send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=True)
    except Exception:
        pass


# ============================================================
# Auth Views
# ============================================================

@transaction.atomic
def register_view(request):
    form = RegistrationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        email = form.cleaned_data['email']
        user = User.objects.create_user(
            username=form.cleaned_data['username'],
            email=email,
            password=form.cleaned_data['password1']
        )
        profile, created = UserProfile.objects.get_or_create(user=user)
        profile.full_name = form.cleaned_data['full_name']
        role = form.cleaned_data['role']
        profile.role = role

        # Multi-Organisation Logic
        if role == 'admin':
            org_name = form.cleaned_data.get('organisation_name')
            if org_name:
                org = Organisation.objects.create(name=org_name, admin_user=user)
                profile.organisation = org
            profile.is_approved = True
        elif role == 'field_worker':
            org = form.cleaned_data.get('organisation_choice')
            if org:
                profile.organisation = org
                # Field workers require approval from their NGO admin
                profile.is_approved = False
                # Notify the NGO admin about the new application
                try:
                    admin_user = org.admin_user
                    send_notification(
                        admin_user,
                        f"📋 New field worker application: {form.cleaned_data['full_name']} wants to join {org.name}.",
                        link='/manage-applications/'
                    )
                except Exception:
                    pass
        else:  # volunteer
            profile.is_approved = True

        profile.locality = form.cleaned_data.get('locality')
        profile.category = form.cleaned_data.get('category')
        profile.save()

        if role == 'field_worker':
            messages.info(request, "Application submitted! Your account is pending approval from the NGO admin. You'll be notified once approved.")
        else:
            messages.success(request, "Account created successfully! Please log in.")
        return redirect('login')
    return render(request, 'NeedsMeshApp/register.html', {'form': form})


def login_view(request):
    form = LoginForm(request, data=request.POST or None)
    if request.method == 'POST':
        # Standard authentication using username and password
        username_or_email = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        # 1. Try standard case-sensitive authentication
        user = authenticate(request, username=username_or_email, password=password)
        
        # 2. Try case-insensitive lookup if first attempt failed
        if not user:
            try:
                user_obj = User.objects.get(username__iexact=username_or_email)
                user = authenticate(request, username=user_obj.username, password=password)
            except (User.DoesNotExist, User.MultipleObjectsReturned):
                pass
        
        # 3. Try lookup by email (fallback)
        if not user:
            try:
                user_obj = User.objects.get(email__iexact=username_or_email)
                user = authenticate(request, username=user_obj.username, password=password)
            except (User.DoesNotExist, User.MultipleObjectsReturned):
                pass
                
        if user:
            login(request, user)
            messages.success(request, f"Welcome back, {user.profile.full_name}!")
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username, email, or password.")
    return render(request, 'NeedsMeshApp/login.html', {'form': form})


@login_required
def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('login')


# ============================================================
# Profile
# ============================================================

@login_required
def profile_view(request):
    profile = request.user.profile
    form = ProfileForm(request.POST or None, request.FILES or None, instance=profile)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Profile updated successfully.")
        return redirect('profile')
    return render(request, 'NeedsMeshApp/profile.html', {'form': form, 'profile': profile})


# ============================================================
# Dashboard — role-aware
# ============================================================

@login_required
def dashboard_view(request):
    profile = request.user.profile
    context = {'profile': profile}

    # Block unapproved field workers — show a pending approval page
    if profile.role == 'field_worker' and not profile.is_approved:
        return render(request, 'NeedsMeshApp/pending_approval.html', {'profile': profile})

    if request.user.is_superuser:
        # SUPER ADMIN: Sees everything
        all_problems = CommunityProblem.objects.select_related('locality', 'submitted_by', 'organisation').all()
        context['all_problems'] = all_problems
        context['is_super_admin'] = True
        # Also show pending applications count
        context['pending_count'] = UserProfile.objects.filter(role='field_worker', is_approved=False).count()
    elif profile.role == 'admin' and profile.organisation:
        # NGO ADMIN: Sees their own organisation's problems
        all_problems = CommunityProblem.objects.filter(organisation=profile.organisation).select_related('locality', 'submitted_by')
        context['all_problems'] = all_problems
        # Show pending applications count for this org
        context['pending_count'] = UserProfile.objects.filter(
            role='field_worker', is_approved=False, organisation=profile.organisation
        ).count()
        # Field Worker KPI: approved FWs in this org with their submission counts
        fw_profiles = UserProfile.objects.filter(
            role='field_worker', is_approved=True, organisation=profile.organisation
        ).select_related('user').annotate(
            submissions=Count('user__submitted_problems')
        ).order_by('-submissions')
        context['fw_counts'] = fw_profiles
    elif profile.role == 'field_worker' and profile.organisation:
        # FIELD WORKER (approved): Sees their own NGO's problems
        my_problems = CommunityProblem.objects.filter(organisation=profile.organisation).order_by('-created_at')
        context['my_problems'] = my_problems

        # Count submissions per field worker
        fw_counts = (
            CommunityProblem.objects
            .filter(submitted_by__profile__role='field_worker')
            .values('submitted_by__profile__full_name', 'submitted_by__email')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        context['fw_counts'] = fw_counts

        # Auto-reminder: urgent problems with no volunteers after 2 days
        reminders = [p for p in my_problems if p.needs_reminder]
        context['reminders'] = reminders

    elif profile.role == 'volunteer':
        # Needs board summary + assigned tasks
        open_problems = CommunityProblem.objects.filter(status='open').select_related('locality')
        assigned = CommunityProblem.objects.filter(
            selected_volunteers=request.user
        ).exclude(status='killed')
        context['open_problems'] = open_problems
        context['assigned_tasks'] = assigned

    # Unread notifications (provided by context_processor too, but explicit here)
    notifications = Notification.objects.filter(recipient=request.user, is_read=False)
    context['notifications'] = notifications

    return render(request, 'NeedsMeshApp/dashboard.html', context)


# ============================================================
# Notifications
# ============================================================

@login_required
def mark_notification_read(request, pk):
    notif = get_object_or_404(Notification, pk=pk, recipient=request.user)
    notif.is_read = True
    notif.save()
    
    if request.headers.get('HX-Request'):
        unread_count = Notification.objects.filter(recipient=request.user, is_read=False).count()
        # Return empty content for the item itself, plus an OOB update for the bell badge
        response_html = f'<span id="notif-count-badge" hx-swap-oob="innerHTML">{unread_count if unread_count > 0 else ""}</span>'
        return HttpResponse(response_html)
        
    if notif.link:
        return redirect(notif.link)
    return redirect('dashboard')


@login_required
def mark_all_read(request):
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    return redirect('dashboard')


# ============================================================
# Survey / Problem Submission (Field Worker)
# ============================================================

@login_required
@role_required('field_worker', 'admin')
def submit_survey(request):
    survey_form = SurveyForm(request.POST or None)
    proof_form = ProofImageForm(request.POST or None, request.FILES or None)

    if request.method == 'POST':
        if survey_form.is_valid():
            problem = survey_form.save(commit=False)
            problem.submitted_by = request.user
            if request.user.profile.organisation:
                problem.organisation = request.user.profile.organisation
            problem.status = 'open'
            problem.save()

            # Handle multiple proof images
            images = request.FILES.getlist('images')
            for img_file in images:
                proof_img = ProofImage.objects.create(image=img_file)
                problem.proof_images.add(proof_img)

            # Create audit record
            SurveySubmission.objects.create(
                field_worker=request.user,
                problem=problem,
                submitted_data_json={
                    'locality': str(problem.locality),
                    'category': problem.category,
                    'urgency': problem.urgency,
                    'detected_date': str(problem.detected_date),
                }
            )

            # Alert admins if urgency is high
            if problem.urgency >= 8:
                admins = User.objects.filter(profile__role='admin')
                for admin_user in admins:
                    send_notification(
                        admin_user,
                        f"⚠️ High urgency ({problem.urgency}/10) problem submitted in {problem.locality}.",
                        link=f'/problems/{problem.pk}/'
                    )

            messages.success(request, "Survey submitted successfully!")
            return redirect('dashboard')
        else:
            messages.error(request, "Please fix the errors below.")

    return render(request, 'NeedsMeshApp/submit_survey.html', {
        'survey_form': survey_form,
        'proof_form': proof_form,
        'localities': Locality.objects.all(),
        'category_choices': CATEGORY_CHOICES,
        'user_locality': request.user.profile.locality,
    })


# ============================================================
# Community Needs Board (Volunteer)
# ============================================================

@login_required
@role_required('volunteer', 'admin', 'field_worker')
def needs_board(request):
    problems = CommunityProblem.objects.filter(
        status__in=['open', 'assigned']
    ).select_related('locality').prefetch_related('interested_volunteers')

    profile = request.user.profile

    # Role-based visibility
    if profile.role == 'admin' or profile.role == 'field_worker':
        if profile.organisation:
            problems = problems.filter(organisation=profile.organisation)
    elif profile.role == 'volunteer' and profile.category:
        problems = problems.filter(category__icontains=profile.category)

    # Filter out fulfilled problems for volunteers
    if profile.role == 'volunteer':
        from django.db.models import Count, F
        problems = problems.annotate(
            interested_count=Count('interested_volunteers', distinct=True)
        ).exclude(
            volunteers_required__gt=0,
            interested_count__gte=F('volunteers_required')
        )

    # Filter support (manual override)
    locality_filter = request.GET.get('locality', '')
    category_filter = request.GET.get('category', '')
    urgency_filter  = request.GET.get('urgency', '')

    if locality_filter:
        problems = problems.filter(locality__name__icontains=locality_filter)
    if category_filter:
        problems = problems.filter(category__icontains=category_filter)
    if urgency_filter:
        problems = problems.filter(urgency=urgency_filter)

    template = 'NeedsMeshApp/needs_board.html'
    if request.headers.get('HX-Request'):
        template = 'NeedsMeshApp/partials/needs_list.html'

    return render(request, template, {
        'problems': problems,
        'localities': Locality.objects.all(),
        'category_choices': CATEGORY_CHOICES,
        'locality_filter': locality_filter,
        'category_filter': category_filter,
        'urgency_filter': urgency_filter,
        'volunteer_category': profile.category if profile.role == 'volunteer' else '',
    })


@login_required
@role_required('volunteer')
def offer_help(request, pk):
    """Volunteer clicks 'I can help' — adds them to interested_volunteers."""
    problem = get_object_or_404(CommunityProblem, pk=pk)
    # Check for active deployments
    has_active = CommunityProblem.objects.filter(
        selected_volunteers=request.user,
        status__in=['open', 'assigned']
    ).exists()
    
    if has_active:
        messages.error(request, "You are currently deployed to an active mission. You can only assist one mission at a time.")
    else:
        if request.user not in problem.interested_volunteers.all():
            problem.interested_volunteers.add(request.user)
            # Notify admins of the SPECIFIC organisation
            admins = User.objects.filter(profile__role='admin', profile__organisation=problem.organisation)
            for admin_user in admins:
                send_notification(
                    admin_user,
                    f"🙋 {request.user.profile.full_name} offered help for: {problem.short_statement()}",
                    link=f'/problems/{problem.pk}/'
                )
            messages.success(request, "You're registered as interested.")
        else:
            messages.info(request, "You have already offered help.")

    if request.headers.get('HX-Request'):
        # Force a full page reload so flash messages show correctly regardless of where they clicked from
        from django.http import HttpResponse
        response = HttpResponse()
        response['HX-Redirect'] = request.META.get('HTTP_REFERER', '/needs-board/')
        return response

    return redirect('needs_board')


# ============================================================
# Problem Detail
# ============================================================

@login_required
def problem_detail(request, pk):
    problem = get_object_or_404(CommunityProblem, pk=pk)
    profile = request.user.profile

    # Volunteers can only see FULL details if selected and event is finalised
    is_selected = request.user in problem.selected_volunteers.all()
    is_interested = request.user in problem.interested_volunteers.all()
    show_full = (
        profile.role in ('admin', 'field_worker')
        or is_selected
        or is_interested
    )

    # Smart volunteer suggestions (for admin)
    suggested_volunteers = []
    if profile.role == 'admin':
        suggested_volunteers = get_suggested_volunteers(problem)

    return render(request, 'NeedsMeshApp/problem_detail.html', {
        'problem': problem,
        'show_full': show_full,
        'is_selected': is_selected,
        'suggested_volunteers': suggested_volunteers,
        'profile': profile,
    })


@login_required
def post_problem_message(request, pk):
    problem = get_object_or_404(CommunityProblem, pk=pk)
    
    # Security: only interested/selected volunteers and the specific NGO admin can post
    is_admin = request.user.profile.role == 'admin' and problem.organisation == request.user.profile.organisation
    is_selected = request.user in problem.selected_volunteers.all()
    is_interested = request.user in problem.interested_volunteers.all()
    
    if not (is_admin or is_selected or is_interested):
        return HttpResponse("Unauthorized", status=403)
        
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        if content:
            msg = ProblemMessage.objects.create(
                problem=problem,
                sender=request.user,
                content=content
            )
            
            # Send automated notification to the other party
            if is_admin:
                for vol in problem.interested_volunteers.all():
                    send_notification(
                        vol,
                        f"💬 New message from Admin regarding mission: {problem.locality}",
                        link=f'/problems/{problem.pk}/'
                    )
            else:
                for admin in User.objects.filter(profile__role='admin', profile__organisation=problem.organisation):
                    role_str = "Team Member" if is_selected else "Applicant"
                    send_notification(
                        admin,
                        f"💬 New message from {role_str} {request.user.profile.full_name} for mission: {problem.locality}",
                        link=f'/problems/{problem.pk}/'
                    )
            
            # If HTMX request, render just the new message
            if request.headers.get('HX-Request'):
                return render(request, 'NeedsMeshApp/partials/message_item.html', {'msg': msg})
                
    return redirect('problem_detail', pk=pk)

@login_required
@role_required('admin')
def accept_volunteer(request, pk, vol_id):
    problem = get_object_or_404(CommunityProblem, pk=pk)
    if problem.organisation != request.user.profile.organisation:
        messages.error(request, "Permission denied.")
        return redirect('problem_detail', pk=pk)
        
    volunteer = get_object_or_404(User, pk=vol_id)
    problem.selected_volunteers.add(volunteer)
    
    send_notification(volunteer, f"✅ Your application for {problem.locality} has been accepted!", link=f'/problems/{problem.pk}/')
    messages.success(request, f"{volunteer.profile.full_name} accepted into the mission.")
    return redirect('problem_detail', pk=pk)

@login_required
@role_required('admin')
def reject_volunteer(request, pk, vol_id):
    problem = get_object_or_404(CommunityProblem, pk=pk)
    if problem.organisation != request.user.profile.organisation:
        messages.error(request, "Permission denied.")
        return redirect('problem_detail', pk=pk)
        
    volunteer = get_object_or_404(User, pk=vol_id)
    problem.interested_volunteers.remove(volunteer)
    problem.selected_volunteers.remove(volunteer)
    
    send_notification(volunteer, f"❌ Your application for {problem.locality} was not selected.")
    messages.info(request, f"{volunteer.profile.full_name}'s application was rejected.")
    return redirect('problem_detail', pk=pk)

def get_suggested_volunteers(problem):
    """
    Smart matching:
    1. Volunteers who helped in same category previously.
    2. Volunteers from same locality (simple string match on profile/past tasks).
    """
    categories = problem.get_category_list()
    locality_name = problem.locality.name if problem.locality else ''

    # Volunteers with past completed tasks in same category
    category_matches = User.objects.filter(
        profile__role='volunteer',
        selected_for__category__iregex='|'.join(categories) if categories else 'none',
        selected_for__status='resolved',
    ).distinct()

    # Volunteers who previously helped in same locality
    locality_matches = User.objects.filter(
        profile__role='volunteer',
        selected_for__locality__name__icontains=locality_name,
        selected_for__status='resolved',
    ).distinct()

    # Combine unique
    combined = (category_matches | locality_matches).distinct()
    return combined[:10]  # Top 10 suggestions


@login_required
def set_volunteers_required(request, pk):
    problem = get_object_or_404(CommunityProblem, pk=pk)
    # Check if NGO Admin of the owning org
    if request.user.profile.role != 'admin' or problem.organisation != request.user.profile.organisation:
        messages.error(request, "Permission denied.")
        return redirect('problem_detail', pk=pk)
    
    if request.method == 'POST':
        form = SetVolunteersRequiredForm(request.POST, instance=problem)
        if form.is_valid():
            form.save()
            messages.success(request, f"Volunteer requirement set to {problem.volunteers_required}.")
    
    return redirect('problem_detail', pk=pk)





# ============================================================
# Admin: Finalise Event
# ============================================================

@login_required
@role_required('admin')
def finalise_event(request, pk):
    problem = get_object_or_404(CommunityProblem, pk=pk)
    form = FinaliseEventForm(request.POST or None, instance=problem)

    if request.method == 'POST' and form.is_valid():
        event = form.save(commit=False)
        event.status = 'assigned'
        event.save()
        form.save_m2m()

        # Notify selected volunteers
        for volunteer in event.selected_volunteers.all():
            send_notification(
                volunteer,
                f"📅 The mission in {problem.locality} is scheduled! Event on {problem.final_event_date:%d %b %Y, %H:%M}.",
                link=f'/problems/{problem.pk}/'
            )
            send_email_notification(
                volunteer,
                subject="Mission Scheduled - NeedsMesh",
                body=(
                    f"Hello {volunteer.profile.full_name},\n\n"
                    f"The mission you are participating in has been scheduled:\n{problem.problem_statement}\n\n"
                    f"Event date: {problem.final_event_date:%d %b %Y, %H:%M}\n"
                    f"Locality: {problem.locality}\n\nThank you for your support!"
                )
            )

        messages.success(request, f"Event finalised! {event.selected_volunteers.count()} volunteer(s) notified.")
        return redirect('problem_detail', pk=problem.pk)

    return render(request, 'NeedsMeshApp/finalise_event.html', {
        'form': form,
        'problem': problem,
    })


# ============================================================
# Admin: Resolve / Kill Problem
# ============================================================

@login_required
@role_required('admin')
def resolve_problem(request, pk):
    problem = get_object_or_404(CommunityProblem, pk=pk)
    form = ResolveProblemForm(request.POST or None, instance=problem)

    if request.method == 'POST' and form.is_valid():
        resolved = form.save(commit=False)
        # Update streak for selected volunteers
        if resolved.status == 'resolved':
            for vol in problem.selected_volunteers.all():
                vol_profile = vol.profile
                vol_profile.completed_tasks += 1
                vol_profile.save()
                if vol_profile.completed_tasks >= 3 and not vol_profile.has_badge:
                    send_notification(
                        vol,
                        "🏅 Congratulations! You've earned the 3-Task Champion Badge!",
                        link='/profile/'
                    )
        resolved.save()
        messages.success(request, f"Problem marked as {resolved.status}.")
        return redirect('problem_detail', pk=problem.pk)

    return render(request, 'NeedsMeshApp/resolve_problem.html', {
        'form': form, 'problem': problem
    })


@login_required
@role_required('admin')
def kill_problem(request, pk):
    problem = get_object_or_404(CommunityProblem, pk=pk)
    form = KillProblemForm(request.POST or None, instance=problem)
    if request.method == 'POST' and form.is_valid():
        p = form.save(commit=False)
        p.status = 'killed'
        p.save()
        messages.warning(request, "Task has been marked as killed.")
        return redirect('dashboard')
    return render(request, 'NeedsMeshApp/kill_problem.html', {'form': form, 'problem': problem})


# ============================================================
# Urgency Heatmap (admin / field_worker)
# ============================================================

@login_required
@role_required('admin', 'field_worker')
def urgency_heatmap(request):
    from django.db.models import Count as DjCount
    localities = Locality.objects.annotate(
        problems_count=DjCount('communityproblems')
    ).all()
    heatmap_data = []
    for loc in localities:
        if loc.latitude and loc.longitude:
            avg_urgency = loc.average_urgency
            if avg_urgency > 0:
                heatmap_data.append({
                    'name': loc.name,
                    'lat': loc.latitude,
                    'lng': loc.longitude,
                    'urgency': avg_urgency,
                    'count': loc.problems_count,
                })

    return render(request, 'NeedsMeshApp/heatmap.html', {
        'heatmap_data_json': json.dumps(heatmap_data),
        'localities': localities,
    })


# ============================================================
# Excel Export (admin only)
# ============================================================

@login_required
@role_required('admin')
def export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        return redirect('dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Community Problems"

    headers = [
        'ID', 'Locality', 'Category', 'Problem Statement', 'Urgency',
        'Status', 'Submitted By', 'Detected Date', 'Final Event Date',
        'Before/After', 'Total Need', 'Community Reaction',
        'Selected Volunteers', 'Created At'
    ]

    # Style header row
    header_fill = PatternFill("solid", fgColor="1a73e8")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    problems = CommunityProblem.objects.select_related('locality', 'submitted_by').all()
    for row_num, p in enumerate(problems, 2):
        selected_names = ', '.join(
            [v.profile.full_name for v in p.selected_volunteers.all()]
        )
        ws.append([
            p.pk,
            str(p.locality) if p.locality else '',
            p.category,
            p.problem_statement,
            p.urgency,
            p.status,
            p.submitted_by.profile.full_name if p.submitted_by else '',
            str(p.detected_date),
            str(p.final_event_date) if p.final_event_date else '',
            p.before_after,
            p.total_need,
            p.community_reaction,
            selected_names,
            str(p.created_at.strftime('%Y-%m-%d %H:%M')),
        ])

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="needsmesh_problems.xlsx"'
    wb.save(response)
    return response


# ============================================================
# Locality Management (Admin)
# ============================================================

@login_required
@role_required('admin')
def manage_localities(request):
    form = LocalityForm(request.POST or None)
    csv_form = CSVImportForm()

    if request.method == 'POST':
        if 'add_locality' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "Locality added.")
            return redirect('manage_localities')

        elif 'upload_csv' in request.POST:
            csv_form = CSVImportForm(request.POST, request.FILES)
            if csv_form.is_valid():
                csv_file = request.FILES['csv_file']
                decoded = csv_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded))
                count = 0
                for row in reader:
                    name = row.get('name', '').strip()
                    if name:
                        lat = row.get('latitude', None) or None
                        lng = row.get('longitude', None) or None
                        Locality.objects.get_or_create(
                            name=name,
                            defaults={'latitude': lat, 'longitude': lng}
                        )
                        count += 1
                messages.success(request, f"{count} localities imported from CSV.")
                return redirect('manage_localities')

    localities = Locality.objects.all()
    return render(request, 'NeedsMeshApp/manage_localities.html', {
        'form': form,
        'csv_form': csv_form,
        'localities': localities,
    })


@login_required
@role_required('admin')
def delete_locality(request, pk):
    locality = get_object_or_404(Locality, pk=pk)
    locality.delete()
    messages.success(request, "Locality deleted.")
    return redirect('manage_localities')


# ============================================================
# Admin: Assign Role to User
# ============================================================

@login_required
@role_required('admin')
def manage_users(request):
    profile = request.user.profile
    if request.user.is_superuser:
        profiles = UserProfile.objects.select_related('user').all()
    else:
        # Admins only see and manage staff inside their own NGO
        profiles = UserProfile.objects.filter(organisation=profile.organisation).select_related('user')
        
    if request.method == 'POST':
        profile_id = request.POST.get('profile_id')
        new_role = request.POST.get('role')
        target_profile = get_object_or_404(UserProfile, pk=profile_id)
        
        # Security: Prevent cross-organisation role changing
        if not request.user.is_superuser and target_profile.organisation != profile.organisation:
             messages.error(request, "Permission denied.")
             return redirect('manage_users')

        if new_role in ['admin', 'field_worker', 'volunteer']:
            target_profile.role = new_role
            target_profile.save()
            messages.success(request, f"Role updated for {profile.full_name}.")
        return redirect('manage_users')
    return render(request, 'NeedsMeshApp/manage_users.html', {'profiles': profiles})


# ============================================================
# Field Worker Application Management (Admin)
# ============================================================

@login_required
@role_required('admin')
def manage_applications(request):
    """NGO Admin view to see & approve/reject pending field worker applications."""
    profile = request.user.profile
    if request.user.is_superuser:
        pending = UserProfile.objects.filter(role='field_worker', is_approved=False).select_related('user', 'organisation', 'locality')
    else:
        pending = UserProfile.objects.filter(
            role='field_worker', is_approved=False, organisation=profile.organisation
        ).select_related('user', 'locality')
    return render(request, 'NeedsMeshApp/manage_applications.html', {'pending': pending})


@login_required
@role_required('admin')
def approve_worker(request, pk):
    """Approve a pending field worker's application."""
    worker_profile = get_object_or_404(UserProfile, pk=pk, role='field_worker')
    admin_profile = request.user.profile

    # Ensure the admin belongs to the same org (unless superuser)
    if not request.user.is_superuser and worker_profile.organisation != admin_profile.organisation:
        messages.error(request, "You can only approve workers from your own organisation.")
        return redirect('manage_applications')

    worker_profile.is_approved = True
    worker_profile.save()

    # Notify the worker
    send_notification(
        worker_profile.user,
        f"🎉 Your application to join {worker_profile.organisation} has been approved! You can now access the platform.",
        link='/dashboard/'
    )
    messages.success(request, f"{worker_profile.full_name}'s application has been approved!")
    return redirect('manage_applications')


@login_required
@role_required('admin')
def reject_worker(request, pk):
    """Reject and remove a pending field worker's application."""
    worker_profile = get_object_or_404(UserProfile, pk=pk, role='field_worker')
    admin_profile = request.user.profile

    if not request.user.is_superuser and worker_profile.organisation != admin_profile.organisation:
        messages.error(request, "You can only reject workers from your own organisation.")
        return redirect('manage_applications')

    worker_name = worker_profile.full_name
    # Remove organisation link and delete the user account
    user = worker_profile.user
    user.delete()  # cascade deletes profile
    messages.warning(request, f"{worker_name}'s application has been rejected and account removed.")
    return redirect('manage_applications')


# ============================================================
# Password Reset via OTP
# ============================================================

import random

def forgot_password_view(request):
    """Step 1: User enters email/username to receive an OTP."""
    if request.method == 'POST':
        identifier = request.POST.get('identifier', '').strip()
        user = None

        # Try finding by username first, then email
        try:
            user = User.objects.get(username__iexact=identifier)
        except User.DoesNotExist:
            try:
                user = User.objects.get(email__iexact=identifier)
            except User.DoesNotExist:
                pass

        if user:
            # Invalidate any previous unused OTPs
            PasswordResetOTP.objects.filter(user=user, is_used=False).update(is_used=True)

            # Generate 6-digit OTP
            otp_code = f"{random.randint(100000, 999999)}"
            PasswordResetOTP.objects.create(user=user, otp_code=otp_code)

            # Send OTP via email (console backend in dev — prints to terminal)
            send_email_notification(
                user,
                subject="Password Reset OTP - NeedsMesh",
                body=(
                    f"Hello {user.profile.full_name},\n\n"
                    f"Your OTP for password reset is: {otp_code}\n\n"
                    f"This code expires in 10 minutes.\n"
                    f"If you did not request this, please ignore this message."
                )
            )

            # Store user id in session for the next steps
            request.session['reset_user_id'] = user.pk
            messages.success(request, "OTP has been sent! Check your email (or terminal in dev mode).")
            return redirect('verify_otp')
        else:
            messages.error(request, "No account found with that username or email.")

    return render(request, 'NeedsMeshApp/forgot_password.html')


def verify_otp_view(request):
    """Step 2: User enters the OTP code."""
    user_id = request.session.get('reset_user_id')
    if not user_id:
        messages.error(request, "Session expired. Please start over.")
        return redirect('forgot_password')

    if request.method == 'POST':
        entered_otp = request.POST.get('otp', '').strip()
        try:
            otp_record = PasswordResetOTP.objects.filter(
                user_id=user_id, otp_code=entered_otp, is_used=False
            ).latest('created_at')
        except PasswordResetOTP.DoesNotExist:
            messages.error(request, "Invalid OTP code. Please try again.")
            return render(request, 'NeedsMeshApp/verify_otp.html')

        if otp_record.is_expired:
            otp_record.is_used = True
            otp_record.save()
            messages.error(request, "OTP has expired. Please request a new one.")
            return redirect('forgot_password')

        # OTP is valid — mark as used and proceed
        otp_record.is_used = True
        otp_record.save()
        request.session['otp_verified'] = True
        messages.success(request, "OTP verified! Set your new password.")
        return redirect('reset_password')

    return render(request, 'NeedsMeshApp/verify_otp.html')


def reset_password_view(request):
    """Step 3: Set new password after OTP verification."""
    user_id = request.session.get('reset_user_id')
    otp_verified = request.session.get('otp_verified', False)

    if not user_id or not otp_verified:
        messages.error(request, "Unauthorized access. Please start the reset process.")
        return redirect('forgot_password')

    user = get_object_or_404(User, pk=user_id)

    if request.method == 'POST':
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        if not password1 or len(password1) < 8:
            messages.error(request, "Password must be at least 8 characters long.")
        elif password1 != password2:
            messages.error(request, "Passwords do not match.")
        else:
            user.set_password(password1)
            user.save()
            # Clean up session
            del request.session['reset_user_id']
            del request.session['otp_verified']
            messages.success(request, "Password reset successful! Please log in with your new password.")
            return redirect('login')

    return render(request, 'NeedsMeshApp/reset_password.html', {'reset_user': user})


# ============================================================
# Home / Landing
# ============================================================

def home_view(request):
    return redirect('login')
